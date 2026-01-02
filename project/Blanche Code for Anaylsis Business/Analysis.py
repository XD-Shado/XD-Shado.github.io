from typing import final

import pandas as pd
import openpyxl
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook

lf = pd.read_excel("ALL_SERVICES_COMBINED.xlsx")

Business22_23 = pd.read_excel(r"C:\Users\aaron\Blanche Code for Anaylsis Business\Business 22-23.xlsx")
Business23_24 = pd.read_excel(r"C:\Users\aaron\Blanche Code for Anaylsis Business\Business 23-24.xlsx")
Business24_25 = pd.read_excel(r"C:\Users\aaron\Blanche Code for Anaylsis Business\Business 24-25.xlsx")

Business22_23["Year"] = 2023
Business23_24["Year"] = 2024
Business24_25["Year"] = 2025

df = pd.concat([Business22_23,Business23_24, Business24_25], ignore_index=True)

combined = pd.concat([Business22_23, Business23_24, Business24_25], ignore_index=True)
final_data = combined[['Category', 'Service ID', 'Description', 'Total', "Year"]] \
              .sort_values(['Category', 'Total'], ascending=[True, False])


final_data.to_excel('ALL_SERVICES_COMBINED.xlsx', index=False)


Final_dataFrame = pd.DataFrame(final_data)

# (₹) the currency they use

total_hair_revenue = lf[lf["Category"] == "Hair"]["Total"].sum()
total_threading_revenue = lf[lf["Category"] == "Threading"]["Total"].sum()
total_waxing_revenue = lf[lf["Category"] == "Waxing"]["Total"].sum()
total_nails_revenue = lf[lf["Category"] == "Nails"]["Total"].sum()
total_shampoo_revenue = lf[lf["Category"] == "Shampoo"]["Total"].sum()
total_shave_revenue = lf[lf["Category"] == "Shave"]["Total"].sum()
total_colour_revenue = lf[lf["Category"] == "Colour"]["Total"].sum()
total_flavour_waxing_revenue = lf[lf["Category"] == "Flavoured_Waxing"]["Total"].sum()
total_treatment_revenue = lf[lf["Category"] == "Treatment"]["Total"].sum()
total_hair_styling_revenue = lf[lf["Category"] == "Hair_styling"]["Total"].sum()
total_facials_revenue = lf[lf["Category"] == "Facials"]["Total"].sum()
total_clean_up_revenue = lf[lf["Category"] == "Clean_Up"]["Total"].sum()
total_bleach_revenue = lf[lf["Category"] == "Bleach"]["Total"].sum()
total_de_tan_revenue = lf[lf["Category"] == "De-Tan"]["Total"].sum()
total_hair_spa_revenue = lf[lf["Category"] == "Hair_spa"]["Total"].sum()
total_straightening_revenue = lf[lf["Category"] == "Straightening"]["Total"].sum()
total_advance_facials_revenue = lf[lf["Category"] == "Advanced_Facials"]["Total"].sum()
total_keratin_revenue = lf[lf["Category"] == "Keratin"]["Total"].sum()
total_product_revenue = lf[lf["Category"] == "PRODUCT"]["Total"].sum()
total_highlights_revenue = lf[lf["Category"] == "HighLights"]["Total"].sum()
total_trimming_revenue = lf[lf["Category"] == "Trimming"]["Total"].sum()
total_makeup_revenue = lf[lf["Category"] == "Makeup"]["Total"].sum()
total_bridal_revenue = lf[lf["Category"] == "Bridal"]["Total"].sum()
total_ear_piercing_revenue = lf[lf["Category"] == "Ear_Piercing"]["Total"].sum()
total_service_revenue = lf[lf["Category"] == "SERVICES"]["Total"].sum()
total_spa_services_revenue = lf[lf["Category"] == "Spa_Services"]["Total"].sum()
total_power_mask_revenue = lf[lf["Category"] == "Power_Mask"]["Total"].sum()





print("The max is ₹",final_data["Total"].max())
print("The min is ₹",final_data["Total"].min())

wb = load_workbook("ALL_SERVICES_COMBINED.xlsx")
ws = wb.active

max_value = f"₹{final_data["Total"].max():,}"
min_value =f"₹{final_data["Total"].min():,}"

max_value_index = final_data.loc[final_data["Total"].idxmax()]
min_value_index = final_data.loc[final_data["Total"].idxmin()]



ws.cell(row=1, column=6).value="MAX"
ws.cell(row=1,column=8).value="MIN"

ws.cell(row=2, column=6).value= max_value
ws.cell(row=2, column=8).value= min_value

wb.save("ALL_SERVICES_COMBINED.xlsx")



#where the results will go
results = []


# loop through all the services provided
services = lf["Category"].unique()
print(services)

num_service = lf["Category"].value_counts()
print(num_service)

#for category in categories:
 #   category_lf = lf[lf["Category"] == category].sort_values("Year")




perdict_year = input("Please type in the year you want perdicted:")

print(perdict_year)

user_perdicted_year = perdict_year

for category in services:
    category_lf = lf[lf["Category"] == category].sort_values("Year")

# x will be category
# y will be sales
    x = category_lf[["Year"]]
    y = category_lf["Total"]

#skip if not enough data to perdict
    if len(x) < 2:
        continue

    model = LinearRegression()
    model.fit(x,y)

    predicted_revenue = model.predict([int[user_perdicted_year]])[0]

    #store results
    results.append({
        "Category": category,
        "Predicted_revenue": predicted_revenue
    })

results_lf = pd.DataFrame(results)

results_lf = results_lf.sort_values("Predicted_revenue", ascending=False)

print("Predicted Revenues for Year", [perdict_year])
print(results_lf)

best_category = results_lf.iloc[0]

print("\nThe service predicted to have the greatest revenue in", perdict_year, "is:", best_category["Category"])


#this is your model creating one
model = LinearRegression()
















